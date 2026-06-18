import os
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

INPUT_CSV = "../data/multi_asset_stress_matrix.csv"
INPUT_DIST = "../data/simulation_distribution.csv"
OUTPUT = "../outputs/Energy_Commodity_Shock_Model.xlsx"

BASE_BRENT = 79.55
BASE_GOLD = 4358.90
REVENUE = 100

SECTORS = {
    "Petrochemicals": {
        "type": "brent",
        "baseline_margin": 0.12,
        "raw_material_pct": 0.60,
        "other_opex_pct": 0.28,
        "elasticity": 0.80,
    },
    "Aviation": {
        "type": "brent",
        "baseline_margin": 0.05,
        "fuel_share_opex": 0.40,
        "opex_pct": 0.95,
        "elasticity": 1.00,
    },
    "Logistics": {
        "type": "brent",
        "baseline_margin": 0.08,
        "fuel_share_opex": 0.25,
        "opex_pct": 0.92,
        "elasticity": 1.00,
    },
    "Consumer Electronics / Jewelry": {
        "type": "gold",
        "baseline_margin": 0.10,
        "raw_material_pct": 0.50,
        "other_opex_pct": 0.40,
        "elasticity": 0.90,
    },
}

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
TOTAL_FONT = Font(bold=True, size=11, name="Calibri")
SECTION_FILL = PatternFill("solid", fgColor="D6E4F0")
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
DANGER_FILL = PatternFill("solid", fgColor="FCE4EC")
GOOD_FILL = PatternFill("solid", fgColor="E8F5E9")
SECTION_FONT = Font(bold=True, size=11, color="1F4E79", name="Calibri")
TITLE_FONT = Font(bold=True, size=16, color="1F4E79", name="Calibri")
SUBTITLE_FONT = Font(size=10, italic=True, color="666666", name="Calibri")

PCT_FMT = '0.0%'
USD_FMT = '$#,##0.00'
NUM_FMT = '#,##0.00'

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
BOTTOM_BORDER = Border(
    bottom=Side(style="medium"),
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),
)


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
        cell.border = THIN_BORDER


def style_border(ws, row, ncols):
    for c in range(1, ncols + 1):
        ws.cell(row=row, column=c).border = THIN_BORDER


def margin_color(v, baseline=0.12):
    if v < 0:
        return Font(bold=True, color="CC0000", name="Calibri")
    if v < baseline * 0.5:
        return Font(bold=True, color="E67E22", name="Calibri")
    return TOTAL_FONT


def compute_margin(sector_type, params, pct_chg, baseline):
    if pct_chg <= 0:
        return baseline
    if sector_type == "brent":
        if "fuel_share_opex" in params:
            opex = params["opex_pct"]
            fuel_pct = opex * params["fuel_share_opex"]
            other_pct = opex - fuel_pct
            new_fuel = fuel_pct * (1 + pct_chg * params["elasticity"])
            return 1 - new_fuel - other_pct
        else:
            raw = params["raw_material_pct"]
            other = params["other_opex_pct"]
            new_raw = raw * (1 + pct_chg * params["elasticity"])
            return 1 - new_raw - other
    else:
        raw = params["raw_material_pct"]
        other = params["other_opex_pct"]
        new_raw = raw * (1 + pct_chg * params["elasticity"])
        return 1 - new_raw - other


def compute_profit_waterfall(sector_type, params, pct_chg, baseline):
    """Return dict of waterfall components."""
    revenue = REVENUE
    if pct_chg <= 0:
        return {"Revenue": revenue, "Cost_of_Goods": 0, "Opex": 0, "Fuel_Cost": 0,
                "Total_Costs": 0, "Profit": revenue * baseline, "Margin": baseline}

    if sector_type == "brent":
        if "fuel_share_opex" in params:
            total_opex_pct = params["opex_pct"]
            fuel_share = params["fuel_share_opex"]
            fuel_pct = total_opex_pct * fuel_share
            other_pct = total_opex_pct - fuel_pct
            fuel_cost = revenue * fuel_pct * (1 + pct_chg * params["elasticity"])
            other_cost = revenue * other_pct
            total_costs = fuel_cost + other_cost
            profit = revenue - total_costs
            margin = profit / revenue
            return {"Revenue": revenue, "Fuel_Cost": fuel_cost, "Other_Opex": other_cost,
                    "Total_Costs": total_costs, "Profit": profit, "Margin": margin}
        else:
            raw_pct = params["raw_material_pct"]
            other_pct = params["other_opex_pct"]
            raw_cost = revenue * raw_pct * (1 + pct_chg * params["elasticity"])
            other_cost = revenue * other_pct
            total_costs = raw_cost + other_cost
            profit = revenue - total_costs
            margin = profit / revenue
            return {"Revenue": revenue, "Raw_Material": raw_cost, "Other_Opex": other_cost,
                    "Total_Costs": total_costs, "Profit": profit, "Margin": margin}
    else:
        raw_pct = params["raw_material_pct"]
        other_pct = params["other_opex_pct"]
        raw_cost = revenue * raw_pct * (1 + pct_chg * params["elasticity"])
        other_cost = revenue * other_pct
        total_costs = raw_cost + other_cost
        profit = revenue - total_costs
        margin = profit / revenue
        return {"Revenue": revenue, "Raw_Material": raw_cost, "Other_Opex": other_cost,
                "Total_Costs": total_costs, "Profit": profit, "Margin": margin}


def main():
    csv = pd.read_csv(INPUT_CSV)
    dist = pd.read_csv(INPUT_DIST) if os.path.exists(INPUT_DIST) else pd.DataFrame()

    prices = {}
    for _, row in csv.iterrows():
        sc_key = row["Scenario"][0]
        asset = row["Asset"]
        prices[(sc_key, asset)] = {
            "median": row["Median_Price"],
            "p5": row["P5_Price"],
            "p95": row["P95_Price"],
            "median_ret": row.get("Median_Return", 0),
            "p5_ret": row.get("P5_Return", 0),
            "p95_ret": row.get("P95_Return", 0),
            "VaR_95%": row.get("VaR_95%", 0),
            "CVaR_95%": row.get("CVaR_95%", 0),
            "VaR_99%": row.get("VaR_99%", 0),
            "CVaR_99%": row.get("CVaR_99%", 0),
            "Max_Drawdown": row.get("Max_Drawdown", 0),
        }

    brent_prices = {sc: prices[(sc, "Brent")]["median"] for sc in ["A", "B", "C"]}
    gold_prices = {sc: prices[(sc, "Gold")]["median"] for sc in ["A", "B", "C"]}

    wb = Workbook()

    # ══════════════════════════════════════════════════════════════
    # SHEET 1 — Executive Dashboard
    # ══════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Executive Dashboard"
    ws.sheet_properties.tabColor = "1F4E79"

    ws.cell(row=1, column=1, value="ENERGY & COMMODITY SHOCK FINANCIAL MODEL").font = TITLE_FONT
    ws.merge_cells("A1:G1")
    ws.cell(row=2, column=1, value="Iran-US Conflict Scenario Analysis | Brent & Gold 30-Day Horizon").font = SUBTITLE_FONT
    ws.merge_cells("A2:G2")

    r = 4
    ws.cell(row=r, column=1, value="SCENARIO PRICE INPUTS").font = SECTION_FONT
    ws.merge_cells(f"A{r}:G{r}")
    r += 1

    for c, h in enumerate(["Metric", "Scenario A (Status Quo)",
                           "Scenario B (Mod. Escalation)",
                           "Scenario C (Severe Disruption)"], 1):
        ws.cell(row=r, column=c, value=h)
    style_header(ws, r, 4)
    r += 1

    for label, key, fmt in [
        ("Brent Crude ($/bbl)", brent_prices, USD_FMT),
        ("Gold ($/oz)", gold_prices, USD_FMT),
    ]:
        ws.cell(row=r, column=1, value=label).font = Font(bold=True, name="Calibri")
        for c, sc in enumerate(["A", "B", "C"], 2):
            cell = ws.cell(row=r, column=c, value=key[sc])
            cell.number_format = fmt
            cell.alignment = Alignment(horizontal="center")
        style_border(ws, r, 4)
        r += 1

    for label, key, base in [
        ("Brent Change vs Current", brent_prices, BASE_BRENT),
        ("Gold Change vs Current", gold_prices, BASE_GOLD),
    ]:
        ws.cell(row=r, column=1, value=label).font = Font(bold=True, name="Calibri")
        for c, sc in enumerate(["A", "B", "C"], 2):
            v = key[sc] / base - 1
            cell = ws.cell(row=r, column=c, value=v)
            cell.number_format = PCT_FMT
            cell.alignment = Alignment(horizontal="center")
        style_border(ws, r, 4)
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="SECTOR-LEVEL PROJECTED MARGIN IMPACT").font = SECTION_FONT
    ws.merge_cells(f"A{r}:G{r}")
    r += 1

    sec_headers = ["Sector", "Linkage", "Baseline Margin",
                   "Scenario A", "Scenario B", "Scenario C"]
    for c, h in enumerate(sec_headers, 1):
        ws.cell(row=r, column=c, value=h)
    style_header(ws, r, 6)
    r += 1

    for name, params in SECTORS.items():
        linkage = "Brent -> Costs" if params["type"] == "brent" else "Gold -> Costs"
        baseline = params["baseline_margin"]
        ws.cell(row=r, column=1, value=name).font = Font(bold=True, name="Calibri")
        ws.cell(row=r, column=2, value=linkage).font = Font(color="555555", name="Calibri")
        ws.cell(row=r, column=3, value=baseline).number_format = PCT_FMT
        ws.cell(row=r, column=3).alignment = Alignment(horizontal="center")

        for c, sc in enumerate(["A", "B", "C"], 4):
            pct_chg = (brent_prices[sc] / BASE_BRENT - 1
                       if params["type"] == "brent"
                       else gold_prices[sc] / BASE_GOLD - 1)
            margin = compute_margin(params["type"], params, pct_chg, baseline)
            cell = ws.cell(row=r, column=c, value=margin)
            cell.number_format = PCT_FMT
            cell.alignment = Alignment(horizontal="center")
            cell.font = margin_color(margin, baseline)
            if margin < 0:
                cell.fill = DANGER_FILL
            elif margin < baseline * 0.5:
                cell.fill = WARN_FILL
        style_border(ws, r, 6)
        r += 1

    for col, w in zip(["A", "B", "C", "D", "E", "F"], [32, 18, 16, 20, 20, 20]):
        ws.column_dimensions[col].width = w

    # ══════════════════════════════════════════════════════════════
    # SHEET 2 — Detailed Impact Model
    # ══════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Detailed Impact Model")
    ws2.sheet_properties.tabColor = "2E75B6"
    ws2.cell(row=1, column=1, value="DETAILED SECTOR-LEVEL IMPACT MODEL").font = Font(
        bold=True, size=14, color="1F4E79", name="Calibri")
    ws2.merge_cells("A1:G1")

    r = 3
    for name, params in SECTORS.items():
        baseline = params["baseline_margin"]
        ws2.cell(row=r, column=1, value=name).font = SECTION_FONT
        ws2.cell(row=r, column=1).fill = SECTION_FILL
        ws2.merge_cells(f"A{r}:G{r}")
        for c in range(1, 8):
            ws2.cell(row=r, column=c).fill = SECTION_FILL
        r += 1

        cols = ["Line Item"] + [f"Scenario {sc}" for sc in ["A", "B", "C"]]
        for c, h in enumerate(cols, 1):
            ws2.cell(row=r, column=c, value=h)
        style_header(ws2, r, 4)
        r += 1

        price_key = brent_prices if params["type"] == "brent" else gold_prices
        base_price = BASE_BRENT if params["type"] == "brent" else BASE_GOLD
        price_unit = "$/bbl" if params["type"] == "brent" else "$/oz"
        price_label = "Brent Price" if params["type"] == "brent" else "Gold Price"
        pct_changes = {sc: price_key[sc] / base_price - 1 for sc in ["A", "B", "C"]}

        impact_lines = [
            ("Baseline Profit Margin", baseline, "single", baseline),
            (f"{price_label} ({price_unit})", price_key, "triplet", None),
            ("Price Change vs Current", pct_changes, "triplet", None),
            ("Cost Elasticity Factor", params["elasticity"], "single", params["elasticity"]),
            ("Implied Cost Change", None, "triplet", None),
            ("Margin Impact (percentage points)", None, "triplet", None),
            ("Projected Profit Margin", None, "triplet", None),
        ]
        for label, val, kind, _ in impact_lines:
            ws2.cell(row=r, column=1, value=label).font = Font(name="Calibri")
            if kind == "single":
                c = 2
                cell = ws2.cell(row=r, column=c, value=val)
                if isinstance(val, float):
                    cell.number_format = PCT_FMT if abs(val) < 1 else NUM_FMT
                ws2.merge_cells(f"B{r}:D{r}")
                cell.alignment = Alignment(horizontal="center")
            elif kind == "triplet":
                for c, sc in enumerate(["A", "B", "C"], 2):
                    if label == "Projected Profit Margin":
                        m = compute_margin(params["type"], params, pct_changes[sc], baseline)
                        cell = ws2.cell(row=r, column=c, value=m)
                        cell.number_format = PCT_FMT
                        cell.font = Font(bold=True, size=11, name="Calibri")
                        if m < 0:
                            cell.font = Font(bold=True, color="CC0000", size=11, name="Calibri")
                    elif label == "Margin Impact (percentage points)":
                        proj = compute_margin(params["type"], params, pct_changes[sc], baseline)
                        cell = ws2.cell(row=r, column=c, value=proj - baseline)
                        cell.number_format = PCT_FMT
                    elif label == "Implied Cost Change":
                        cell = ws2.cell(row=r, column=c, value=pct_changes[sc] * params["elasticity"])
                        cell.number_format = PCT_FMT
                    elif label == f"{price_label} ({price_unit})":
                        cell = ws2.cell(row=r, column=c, value=price_key[sc])
                        cell.number_format = USD_FMT
                    elif label == "Price Change vs Current":
                        cell = ws2.cell(row=r, column=c, value=pct_changes[sc])
                        cell.number_format = PCT_FMT
                    cell.alignment = Alignment(horizontal="center")
            style_border(ws2, r, 4)
            r += 1
        r += 1

    ws2.column_dimensions["A"].width = 36
    for col in ["B", "C", "D"]:
        ws2.column_dimensions[col].width = 22

    # ══════════════════════════════════════════════════════════════
    # SHEET 3 — Assumptions Register
    # ══════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Assumptions Register")
    ws3.sheet_properties.tabColor = "70AD47"
    ws3.cell(row=1, column=1, value="MODEL ASSUMPTIONS & PARAMETERS").font = TITLE_FONT
    ws3.merge_cells("A1:D1")
    ws3.cell(row=2, column=1, value="Baseline Brent Price").font = Font(bold=True, name="Calibri")
    ws3.cell(row=2, column=2, value=BASE_BRENT).number_format = USD_FMT
    ws3.cell(row=3, column=1, value="Baseline Gold Price").font = Font(bold=True, name="Calibri")
    ws3.cell(row=3, column=2, value=BASE_GOLD).number_format = USD_FMT
    ws3.cell(row=4, column=1, value="Normalized Revenue per Sector").font = Font(bold=True, name="Calibri")
    ws3.cell(row=4, column=2, value=f"${REVENUE} (numeraire)")

    r = 6
    param_headers = ["Sector", "Parameter", "Value", "Notes"]
    for c, h in enumerate(param_headers, 1):
        ws3.cell(row=r, column=c, value=h)
    style_header(ws3, r, 4)
    r += 1

    for name, params in SECTORS.items():
        notes_map = {
            "Baseline Profit Margin": "Industry-average EBIT margin estimate",
            "Elasticity Factor": ("0.8x Brent pass-through" if params["type"] == "brent"
                                  else "0.9x Gold pass-through"),
        }
        for param, val in [
            ("Baseline Profit Margin", params["baseline_margin"]),
            ("Elasticity Factor", params["elasticity"]),
        ]:
            ws3.cell(row=r, column=1, value=name).font = Font(bold=True, name="Calibri")
            ws3.cell(row=r, column=2, value=param)
            cell = ws3.cell(row=r, column=3, value=val)
            cell.number_format = PCT_FMT
            ws3.cell(row=r, column=4, value=notes_map.get(param, "")).font = Font(color="666666", size=9, name="Calibri")
            r += 1
        if "raw_material_pct" in params:
            for label, key, note in [
                ("Raw Materials % of Revenue", "raw_material_pct", "Direct commodity-input share"),
                ("Other Opex % of Revenue", "other_opex_pct", "Labor, depreciation, SG&A"),
            ]:
                ws3.cell(row=r, column=1, value=name).font = Font(bold=True, name="Calibri")
                ws3.cell(row=r, column=2, value=label)
                ws3.cell(row=r, column=3, value=params[key]).number_format = PCT_FMT
                ws3.cell(row=r, column=4, value=note).font = Font(color="666666", size=9, name="Calibri")
                r += 1
        if "fuel_share_opex" in params:
            for label, key, note in [
                ("Fuel % of Operating Expenses", "fuel_share_opex", "Jet fuel / diesel as share of total opex"),
                ("Total Opex % of Revenue", "opex_pct", "Industry cost-to-revenue ratio"),
            ]:
                ws3.cell(row=r, column=1, value=name).font = Font(bold=True, name="Calibri")
                ws3.cell(row=r, column=2, value=label)
                ws3.cell(row=r, column=3, value=params[key]).number_format = PCT_FMT
                ws3.cell(row=r, column=4, value=note).font = Font(color="666666", size=9, name="Calibri")
                r += 1

    ws3.column_dimensions["A"].width = 34
    ws3.column_dimensions["B"].width = 32
    ws3.column_dimensions["C"].width = 16
    ws3.column_dimensions["D"].width = 40

    # ══════════════════════════════════════════════════════════════
    # SHEET 4 — Sensitivity Analysis
    # ══════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Sensitivity Analysis")
    ws4.sheet_properties.tabColor = "C00000"
    ws4.cell(row=1, column=1, value="SENSITIVITY ANALYSIS — PRICE SHOCK vs MARGIN IMPACT").font = TITLE_FONT
    ws4.merge_cells("A1:H1")
    ws4.cell(row=2, column=1, value="Each cell shows projected margin under a given price shock (columns)").font = SUBTITLE_FONT
    ws4.merge_cells("A2:H2")

    shock_grid = np.arange(-0.1, 0.71, 0.1)
    r = 4
    ws4.cell(row=r, column=1, value="Sector").font = SECTION_FONT
    for c, shock in enumerate(shock_grid, 2):
        cell = ws4.cell(row=r, column=c, value=shock)
        cell.number_format = PCT_FMT
        cell.alignment = Alignment(horizontal="center")
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
    ws4.cell(row=r, column=1).font = HEADER_FONT
    ws4.cell(row=r, column=1).fill = HEADER_FILL
    ws4.cell(row=r, column=1).border = THIN_BORDER
    r += 1

    for name, params in SECTORS.items():
        baseline = params["baseline_margin"]
        ws4.cell(row=r, column=1, value=name).font = Font(bold=True, name="Calibri")
        for c, shock in enumerate(shock_grid, 2):
            pct_chg = max(0, shock)
            margin = compute_margin(params["type"], params, pct_chg, baseline)
            cell = ws4.cell(row=r, column=c, value=margin)
            cell.number_format = PCT_FMT
            cell.alignment = Alignment(horizontal="center")
            cell.font = margin_color(margin, baseline)
            if margin < 0:
                cell.fill = DANGER_FILL
            elif margin < baseline * 0.5:
                cell.fill = WARN_FILL
            elif margin >= baseline:
                cell.fill = GOOD_FILL
            cell.border = THIN_BORDER
        r += 1

    ws4.column_dimensions["A"].width = 34
    for c in range(2, 11):
        ws4.column_dimensions[get_column_letter(c)].width = 14

    # ══════════════════════════════════════════════════════════════
    # SHEET 5 — Risk Metrics
    # ══════════════════════════════════════════════════════════════
    ws5 = wb.create_sheet("Risk Metrics")
    ws5.sheet_properties.tabColor = "7030A0"
    ws5.cell(row=1, column=1, value="RISK METRICS — VaR, CVaR & TAIL RISK ANALYSIS").font = TITLE_FONT
    ws5.merge_cells("A1:G1")

    r = 3
    ws5.cell(row=r, column=1, value="Metric").font = HEADER_FONT
    ws5.cell(row=r, column=1).fill = HEADER_FILL
    ws5.cell(row=r, column=1).border = THIN_BORDER

    risk_metrics = ["VaR_90%", "CVaR_90%", "VaR_95%", "CVaR_95%", "VaR_99%", "CVaR_99%", "Max_Drawdown", "Max_Upside"]
    col_idx = 2
    for sc_label in ["A: Status Quo (State 0)", "B: Moderate Escalation", "C: Severe Disruption"]:
        for asset in ["Brent", "Gold"]:
            cell = ws5.cell(row=r, column=col_idx, value=f"{asset}\n{sc_label}")
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
            cell.border = THIN_BORDER
            col_idx += 1

    col_idx = 2
    for sc_label in ["A: Status Quo (State 0)", "B: Moderate Escalation", "C: Severe Disruption"]:
        for _ in ["Brent", "Gold"]:
            ws5.column_dimensions[get_column_letter(col_idx)].width = 18
            col_idx += 1

    r = 4
    for metric in risk_metrics:
        ws5.cell(row=r, column=1, value=metric).font = Font(bold=True, name="Calibri")
        ws5.cell(row=r, column=1).border = THIN_BORDER
        col_idx = 2
        for sc_key, sc_label in [("A", "A: Status Quo (State 0)"),
                                  ("B", "B: Moderate Escalation"),
                                  ("C", "C: Severe Disruption")]:
            for asset in ["Brent", "Gold"]:
                val = prices[(sc_key, asset)].get(metric, None)
                if val is not None:
                    cell = ws5.cell(row=r, column=col_idx, value=val)
                    cell.number_format = PCT_FMT
                    cell.alignment = Alignment(horizontal="center")
                    if "Max" in metric:
                        cell.font = margin_color(val, 0.0) if val < 0 else Font(bold=True, color="006400", name="Calibri")
                    elif "VaR" in metric or "CVaR" in metric:
                        cell.font = margin_color(val, 0.0)
                        if val < -0.3:
                            cell.fill = DANGER_FILL
                        elif val < -0.1:
                            cell.fill = WARN_FILL
                cell.border = THIN_BORDER
                col_idx += 1
        r += 1

    # ══════════════════════════════════════════════════════════════
    # SHEET 6 — Profit Waterfall (Severe Disruption)
    # ══════════════════════════════════════════════════════════════
    ws6 = wb.create_sheet("Profit Waterfall - Severe")
    ws6.sheet_properties.tabColor = "FF6600"
    ws6.cell(row=1, column=1, value="PROFIT WATERFALL — SEVERE DISRUPTION SCENARIO").font = TITLE_FONT
    ws6.merge_cells("A1:G1")
    ws6.cell(row=2, column=1, value="Per $100 revenue: how costs erode profit margin").font = SUBTITLE_FONT
    ws6.merge_cells("A2:G2")

    sc_key = "C"
    r = 4
    for name, params in SECTORS.items():
        pct_chg = (brent_prices[sc_key] / BASE_BRENT - 1
                   if params["type"] == "brent"
                   else gold_prices[sc_key] / BASE_GOLD - 1)
        waterfall = compute_profit_waterfall(params["type"], params, pct_chg, params["baseline_margin"])

        ws6.cell(row=r, column=1, value=name).font = SECTION_FONT
        ws6.cell(row=r, column=1).fill = SECTION_FILL
        ws6.merge_cells(f"A{r}:G{r}")
        for c in range(1, 8):
            ws6.cell(row=r, column=c).fill = SECTION_FILL
        r += 1

        wf_headers = ["Component", "Value ($)", "% of Revenue"]
        for c, h in enumerate(wf_headers, 1):
            ws6.cell(row=r, column=c, value=h)
        style_header(ws6, r, 3)
        r += 1

        components = []
        components.append(("Revenue", waterfall["Revenue"], 1.0))
        if "Raw_Material" in waterfall:
            components.append(("Raw Material Cost", -waterfall["Raw_Material"],
                               -waterfall["Raw_Material"] / waterfall["Revenue"]))
        if "Fuel_Cost" in waterfall:
            components.append(("Fuel Cost", -waterfall["Fuel_Cost"],
                               -waterfall["Fuel_Cost"] / waterfall["Revenue"]))
        if "Other_Opex" in waterfall:
            components.append(("Other Operating Costs", -waterfall["Other_Opex"],
                               -waterfall["Other_Opex"] / waterfall["Revenue"]))
        components.append(("= Profit", waterfall["Profit"], waterfall["Margin"]))

        for label, val, pct in components:
            ws6.cell(row=r, column=1, value=label).font = Font(name="Calibri",
                      bold=(label == "= Profit" or label == "Revenue"))
            cell = ws6.cell(row=r, column=2, value=val)
            cell.number_format = USD_FMT
            cell.alignment = Alignment(horizontal="center")
            cell = ws6.cell(row=r, column=3, value=pct)
            cell.number_format = PCT_FMT
            cell.alignment = Alignment(horizontal="center")

            if val < 0:
                cell.fill = DANGER_FILL
            elif label == "= Profit" and pct < 0:
                ws6.cell(row=r, column=2).font = Font(bold=True, color="CC0000", name="Calibri")
                ws6.cell(row=r, column=3).font = Font(bold=True, color="CC0000", name="Calibri")
                ws6.cell(row=r, column=2).fill = DANGER_FILL
                ws6.cell(row=r, column=3).fill = DANGER_FILL
            elif label == "= Profit" and pct > 0:
                ws6.cell(row=r, column=2).fill = GOOD_FILL
                ws6.cell(row=r, column=3).fill = GOOD_FILL
            style_border(ws6, r, 3)
            r += 1
        r += 1

    ws6.column_dimensions["A"].width = 28
    ws6.column_dimensions["B"].width = 18
    ws6.column_dimensions["C"].width = 18

    wb.save(OUTPUT)
    print(f"Workbook saved: {OUTPUT}  (6 sheets)")


if __name__ == "__main__":
    main()
